import openpyxl

def convert_xlsx(file_path):
    wb=openpyxl.load_workbook(file_path)
    data=[]
    for s in wb.sheetnames:
        ws=wb[s]
        for row in ws.iter_rows(values_only=True):
            data.append(" | ".join(str(c) if c else "" for c in row))
    return "\n".join(data)
